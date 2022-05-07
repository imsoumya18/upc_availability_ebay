import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import os

upc_list = []
upc_set = ()
print(os.getcwd() + '\\sheets')
for file in os.listdir(os.getcwd() + '\\sheets'):
    wb = openpyxl.load_workbook(os.getcwd() + '\\sheets\\' + file)
    wb2 = openpyxl.Workbook()

    ws = wb.worksheets[0]
    ws2 = wb2.create_sheet(title='Sheet1', index=0)

    i = 1
    while ws.cell(i, 1).value is not None:
        if ws.cell(i, 3).value == 'SOLD':
            ws.delete_rows(i, 1)
        i += 1

    i = 1
    while ws.cell(i, 1).value is not None:
        upc_list.append(ws.cell(i, 1).value)
        i += 1

upc_set = set(upc_list)

redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')

ws2.cell(1, 1).value = 'UPC'
ws2.cell(1, 1).font = Font(size=20)
ws2.cell(1, 1).alignment = Alignment(horizontal='center')
ws2.cell(1, 1).fill = redFill
ws2.cell(1, 2).value = 'Count'
ws2.cell(1, 2).font = Font(size=20)
ws2.cell(1, 2).alignment = Alignment(horizontal='center')
ws2.cell(1, 2).fill = redFill
ws2.cell(1, 3).value = 'Boxes'
ws2.cell(1, 3).font = Font(size=20)
ws2.cell(1, 3).alignment = Alignment(horizontal='center')
ws2.cell(1, 3).fill = redFill

j = 2
for i in upc_set:
    box = set()
    ws2.cell(j, 1).value = i
    ws2.cell(j, 1).number_format = '0'
    if upc_list.count(i) != 1:
        ws2.cell(j, 2).value = 'x' + str(upc_list.count(i))
        ws2.cell(j, 2).alignment = Alignment(horizontal='right')
    k = 1
    while ws.cell(k, 1).value is not None:
        if ws.cell(k, 1).value == i:
            box.add(ws.cell(k, 2).value)
        k += 1
    ws2.cell(j, 3).value = ', '.join(sorted(box))
    ws2.cell(j, 3).alignment = Alignment(wrap_text=True, horizontal='right')
    j += 1

ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 16

try:
    wb2.save('output.xlsx')
    print('Successful')
except:
    print('Not successful')
